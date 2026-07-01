"""
processar_dados.py v6 — Dashboard Analítica de CX · Xtay
Módulo reutilizável: processar(rows_droz, rows_occ) → payload dict
Uso local: python3 processar_dados.py  (lê Base de Atendimento.xlsx)
"""
import json, re
from collections import defaultdict, Counter
from datetime import datetime, timedelta, timezone

# GMT-3 (Horário de Brasília) — item 12
TZ_BR = timezone(timedelta(hours=-3))

ARQUIVO_ENTRADA  = "Base de Atendimento.xlsx"
ARQUIVO_TEMPLATE = "template.html"
ARQUIVO_SAIDA    = "dashboard_cx.html"

TAGS_STATUS_CLIENTE = {
    "esta_hospedado","nao_esta_hospedado",
    "nao_possui_checkin_hoje_nem_futuros","existe_checkin_futuro",
}
MAPA_PRIORIDADE = {"prioridade_alta":"Alta","prioridade_media":"Média","prioridade_baixa":"Baixa"}

# Empreendimentos ativos (lista oficial — item 4)
EMPREEND_ATIVOS = {
    "Linked Batel","PUC","Cais","Atrium","Upper West",
    "Vila Madalena","Fuji","Oslo","Campos Sales","Simple Smart","Soho","Ipiranga"
}

EMPREEND_KEYWORDS = [
    ("linked batel","Linked Batel"),("linked_batel","Linked Batel"),
    ("campos salles","Campos Sales"),("campos_salles","Campos Sales"),
    ("campos sales","Campos Sales"),("campos_sales","Campos Sales"),
    ("upper west","Upper West"),("upper_west","Upper West"),
    ("simple smart","Simple Smart"),("simple_smart","Simple Smart"),
    ("simple smarrt","Simple Smart"),("simple smarrt","Simple Smart"),
    ("vila madalena","Vila Madalena"),("vila_madalena","Vila Madalena"),
    ("the bridge","The Bridge"),("the_bridge","The Bridge"),
    ("the spot one","The Spot One"),("the_spot_one","The Spot One"),("tso","The Spot One"),
    ("guarulhos","Guarulhos Aeroporto"),
    ("princess","Princess Curitiba"),
    ("bk30 largo do arouche","BK30 Largo do Arouche"),("bk30 santana","BK30 Santana"),
    ("restinga","Restinga"),
    ("ipiranga","Ipiranga"),("ip alto","Ipiranga"),
    # ── Inativos — reconhecidos para filtro, não listados nos ativos ──
    ("gru","Guarulhos Aeroporto"),
    ("oscar freire","Oscar Freire"),("oscar_freire","Oscar Freire"),
    ("berrini","Berrini"),
    ("upper","Upper West"),("simple","Simple Smart"),
    ("cais","Cais"),("atrium","Atrium"),("fuji","Fuji"),
    ("soho","Soho"),("oslo","Oslo"),("puc","PUC"),
]

CLUSTERS = {
    "Vila Madalena":"São Paulo","Upper West":"São Paulo","Ipiranga":"São Paulo",
    "BK30 Largo do Arouche":"São Paulo","BK30 Santana":"São Paulo",
    "Guarulhos Aeroporto":"São Paulo",
    "Cais":"Porto Alegre","Simple Smart":"João Pessoa",
    "Campos Sales":"Curitiba","PUC":"Curitiba",
    "Princess Curitiba":"Curitiba",
    "Linked Batel":"Linked Batel",
    "Atrium":"Santa Catarina","Fuji":"Santa Catarina",
    "Soho":"Santa Catarina","Oslo":"Santa Catarina","The Bridge":"Santa Catarina",
    "The Spot One":"Santa Catarina","Restinga":"Santa Catarina",
}

CLASSIF_EXACT = {
    # ── Recepção Digital ──────────────────────────────────────
    "abertura_remota_check_out_vencido":                    "recepcao_digital",
    "abertura_remota_dificuldades_do_hospede":              "recepcao_digital",
    "aviso_de_check_out":                                   "recepcao_digital",
    "check_in_de_dependente":                               "recepcao_digital",
    "check_in_duvidas":                                     "recepcao_digital",
    "check_in_manual_dificuldades_do_hospede":              "recepcao_digital",
    "checkin_b2b":                                          "recepcao_digital",
    "checkout_duvidas":                                     "recepcao_digital",
    "early_check_in_duvidas":                               "recepcao_digital",
    "inclusao_check_in_de_dependente":                      "recepcao_digital",
    "late_check_out_duvidas":                               "recepcao_digital",
    "solicitacao_de_early_check_in_autorizado_com_custos":  "recepcao_digital",
    "solicitacao_de_early_check_in_cortesia":               "recepcao_digital",
    "solicitacao_de_early_check_in_sem_disponibilidade":    "recepcao_digital",
    "solicitacao_de_late_check_out_autorizado_com_custos":  "recepcao_digital",
    "solicitacao_de_late_check_out_cortesia":               "recepcao_digital",
    "solicitacao_de_late_check_out_sem_disponibilidade":    "recepcao_digital",
    "troca_de_apartamento":                                 "recepcao_digital",
    "cadastro_visitante":                                   "recepcao_digital",
    "duvidas_cadastro_visitante":                           "recepcao_digital",
    "duvidas_gerais_senha_de_acesso":                       "recepcao_digital",
    "estacionamento_dificuldades_de_acesso":                "recepcao_digital",
    "estacionamento_duvidas":                               "recepcao_digital",
    "estacionamento_reserva":                               "recepcao_digital",
    # ── Tech ─────────────────────────────────────────────────
    "check_in_erro_de_verificacao":         "tech",
    "check_in_manual_documento_digital":    "tech",
    "check_in_reserva_nao_localizada":      "tech",
    "checkout_manual":                      "tech",
    "problemas_com_a_facial":               "tech",
    "problemas_com_senha_da_porta":         "tech",
    "problemas_com_link_de_pagamento":      "tech",
    "problemas_com_pagamento":              "tech",
    # ── Gestão de Reservas ───────────────────────────────────
    "alteracao_de_datas":                               "gestao_reservas",
    "alteracao_de_reserva_inclusao_de_acompanhante":    "gestao_reservas",
    "alteracao_de_reserva_inclusao_de_dependente":      "gestao_reservas",
    "alteracao_de_reserva_prorrogacao":                 "gestao_reservas",
    "cancelamento_de_reserva_dentro_da_politica":       "gestao_reservas",
    "cancelamento_de_reserva_fora_da_politica":         "gestao_reservas",
    "cupom_de_desconto_reservas":                       "gestao_reservas",
    "reserva_upgrade":                                  "gestao_reservas",
    "reembolso":                                        "gestao_reservas",
    "suspeita_de_fraude":                               "gestao_reservas",
    "alteracao_de_dados_cadastrais":                    "gestao_reservas",
    # ── Comercial ────────────────────────────────────────────
    "atendimento-comercial":                        "comercial",
    "comercial":                                    "comercial",
    "comercial_hospedar":                           "comercial",
    "comercial_morar":                              "comercial",
    "confirmacao_de_reserva":                       "comercial",
    "cotacao_orcamento_venda_finalizada":            "comercial",
    "cotacao_orcamento_venda_nao_finalizada":        "comercial",
    "duvidas_omnibees":                             ["comercial","duvidas_gerais"],
    "hospedar":                                     "comercial",
    "mensalista_agendamento_de_limpeza_quinzenal":  "comercial",
    "mensalista_contrato_venda":                    "comercial",
    "mensalista_cotacao_e_duvidas":                 "comercial",
    "mensalista_solicitacao_de_link_de_pagamento":  "comercial",
    "morar":                                        "comercial",
    "novos_negocios":                               "comercial",
    "novos_negocios_marketing":                     "comercial",
    # ── Dúvidas Gerais ───────────────────────────────────────
    "informacoes_de_acesso_wi_fi":  "duvidas_gerais",
    "maleiro":                      "duvidas_gerais",
    "detalhes_da_acomodacao":       "duvidas_gerais",
    "duvidas_descarte_lixo":        "duvidas_gerais",
    "duvidas_gerais":               "duvidas_gerais",
    "duvidas_predio":               "duvidas_gerais",   # adicionado v6
    "duvidas_sobre_limpeza":        "duvidas_gerais",
    # ── Serviços Solicitados ─────────────────────────────────
    "solicitacao_de_nota_fiscal":               "servicos",
    "solicitacao_de_servico_item_extra":        "servicos",
    "solicitacao_de_servico_limpeza":           "servicos",
    "solicitacao_de_servico_taxa_pet":          "servicos",
    "solicitacao_de_servico_troca_de_enxoval":  "servicos",
    "solicitacao_de_servico_troca_de_toalhas":  "servicos",
    "reposicao_de_amenidades":                  "servicos",
    "mini_mercado":                             "servicos",
    "lavanderia_omo":                           "servicos",
    "cafe_da_manha":                            "servicos",
    # ── Manutenção ───────────────────────────────────────────
    "solicitacao_de_servico_manutencao":    ["servicos","manutencao"],
    "ar_condicionado":                      "manutencao",
    "hidraulica":                           "manutencao",
    "item_danificado_no_apto":              ["manutencao","operacional"],
    "item_faltante_na_unidade":             ["manutencao","operacional"],
    "aviso_de_falta_de_energia":            ["manutencao","operacional"],
    # ── Problemas com Limpeza ────────────────────────────────
    "problemas_com_limpeza":    "limpeza",
    "reclamacao_de_limpeza":    "limpeza",
    "limpeza_nao_realizada":    "limpeza",
    # ── Gestão Operacional ───────────────────────────────────
    "reclamacao_da_vista":      "operacional",
    "reclamacao_de_barulho":    "operacional",
    "reclamacao_de_internet":   "operacional",
    "reclamacao_estadia":       "operacional",
    "feedback_estadia":         "operacional",
    "achados_e_perdidos":       "operacional",
    "conversa_operacional":     "operacional",
    # ── Marketing e Relacionamento ───────────────────────────
    "mkt_hospede_frequent": "marketing",
    "hospede_mkt":          "marketing",
    "trabalhe_conosco":     "marketing",
    # ── Tech (adicionais) ────────────────────────────────────
    "erro_de_verificacao_checkin_manual": "tech",
    "erro_senha_null":                    "tech",
    # ── Manutenção (adicionais) ──────────────────────────────
    "eletrica":               "manutencao",
    "mobiliario":             "manutencao",
    "equipamentos":           "manutencao",
    "aviso_falta_de_agua":    ["manutencao","operacional"],
    # ── Contato Ativo ────────────────────────────────────────
    "contato_ativo":                        "contato_ativo",
    "contato_ativo_comunicacoes":           "contato_ativo",
    "contato_ativo_cobranca":               "contato_ativo",
    "contato_ativo_autorizacao_entrada":    "contato_ativo",
    # ── Improdutivos ─────────────────────────────────────────
    "falta_de_interacao":   "improdutivos",
    "sem_retorno":          "improdutivos",
    "teste":                "improdutivos",
}

CLASSIF_PREFIXOS = [
    ("mensalista_",                     "comercial"),
    ("cotacao_orcamento_",              "comercial"),
    ("novos_negocios",                  "comercial"),
    ("cancelamento_de_reserva_",        "gestao_reservas"),
    ("alteracao_de_reserva_",           "gestao_reservas"),
    ("solicitacao_de_early_check_in_",  "recepcao_digital"),
    ("solicitacao_de_late_check_out_",  "recepcao_digital"),
    ("abertura_remota_",                "recepcao_digital"),
    ("solicitacao_de_servico_",         "servicos"),
    ("contato_ativo_",                  "contato_ativo"),   # cobre variações futuras
]

CLASSIF_META = {
    "recepcao_digital":  {"label":"Recepção Digital",         "cor":"#2196F3"},
    "tech":              {"label":"Tech",                      "cor":"#9B59B6"},
    "gestao_reservas":   {"label":"Gestão de Reservas",       "cor":"#FF5722"},
    "comercial":         {"label":"Comercial",                 "cor":"#FF6A13"},
    "duvidas_gerais":    {"label":"Dúvidas Gerais",            "cor":"#4F79E4"},
    "servicos":          {"label":"Serviços Solicitados",      "cor":"#00BCD4"},
    "manutencao":        {"label":"Manutenção",                "cor":"#795548"},
    "limpeza":           {"label":"Problemas com Limpeza",     "cor":"#E91E63"},
    "operacional":       {"label":"Gestão Operacional",        "cor":"#E67E22"},
    "marketing":         {"label":"Marketing e Relacionamento","cor":"#27AE60"},
    "contato_ativo":     {"label":"Contato Ativo",             "cor":"#1ABC9C"},
    "improdutivos":      {"label":"Improdutivos",              "cor":"#95A5A6"},
}
CLASSIF_ORDER = [
    "recepcao_digital","tech","gestao_reservas","comercial","duvidas_gerais",
    "servicos","manutencao","limpeza","operacional","marketing","contato_ativo","improdutivos",
]

# Apenas empreendimentos ativos — item 4
EMPREEND_ORDER = [
    "Linked Batel","PUC","Upper West","Vila Madalena","Cais",
    "Atrium","Fuji","Oslo","Campos Sales","Simple Smart","Soho","Ipiranga",
]
CLUSTER_ORDER = ["São Paulo","Curitiba","Linked Batel","Porto Alegre","Santa Catarina","João Pessoa"]

NM = {"01":"Jan","02":"Fev","03":"Mar","04":"Abr","05":"Mai","06":"Jun",
      "07":"Jul","08":"Ago","09":"Set","10":"Out","11":"Nov","12":"Dez"}

MES_NUM = {
    "jan":"01","fev":"02","mar":"03","abr":"04","mai":"05","jun":"06",
    "jul":"07","ago":"08","set":"09","out":"10","nov":"11","dez":"12",
    "janeiro":"01","fevereiro":"02","marco":"03","abril":"04","maio":"05",
    "junho":"06","julho":"07","agosto":"08","setembro":"09","outubro":"10",
    "novembro":"11","dezembro":"12",
}

LABELS = {
    "check_in_duvidas":"Check-in — dúvidas","duvidas_gerais":"Dúvidas gerais",
    "sem_retorno":"Sem retorno","check_in_manual_dificuldades_do_hospede":"Check-in manual — dificuldades",
    "alteracao_de_reserva_prorrogacao":"Prorrogação de reserva",
    "problemas_com_senha_da_porta":"Problemas com senha da porta",
    "solicitacao_de_early_check_in_cortesia":"Early check-in (cortesia)",
    "solicitacao_de_servico_manutencao":"Manutenção",
    "early_check_in_duvidas":"Early check-in — dúvidas",
    "estacionamento_duvidas":"Estacionamento — dúvidas",
    "solicitacao_de_late_check_out_cortesia":"Late check-out (cortesia)",
    "troca_de_apartamento":"Troca de apartamento",
    "cotacao_orcamento_venda_nao_finalizada":"Cotação não finalizada",
    "cotacao_orcamento_venda_finalizada":"Cotação finalizada",
    "inclusao_check_in_de_dependente":"Inclusão dependente no check-in",
    "duvidas_gerais_senha_de_acesso":"Dúvidas senha de acesso",
    "problemas_com_a_facial":"Problemas com facial",
    "mensalista_agendamento_de_limpeza_quinzenal":"Mensalista — limpeza quinzenal",
    "achados_e_perdidos":"Achados e perdidos","solicitacao_de_nota_fiscal":"Nota fiscal",
    "solicitacao_de_servico_item_extra":"Item extra",
    "check_in_reserva_nao_localizada":"Check-in — reserva não localizada",
    "confirmacao_de_reserva":"Confirmação de reserva",
    "solicitacao_de_servico_limpeza":"Solicitação de limpeza",
    "reclamacao_de_limpeza":"Reclamação de limpeza",
    "late_check_out_duvidas":"Late check-out — dúvidas",
    "mensalista_cotacao_e_duvidas":"Mensalista — cotação",
    "check_in_de_dependente":"Check-in de dependente",
    "cafe_da_manha":"Café da manhã","duvidas_sobre_limpeza":"Dúvidas sobre limpeza",
    "duvidas_predio":"Dúvidas sobre o prédio",
    "alteracao_de_datas":"Alteração de datas",
    "cancelamento_de_reserva_fora_da_politica":"Cancelamento fora da política",
    "cancelamento_de_reserva_dentro_da_politica":"Cancelamento dentro da política",
    "solicitacao_de_servico_troca_de_toalhas":"Troca de toalhas",
    "solicitacao_de_early_check_in_autorizado_com_custos":"Early check-in (com custo)",
    "solicitacao_de_late_check_out_autorizado_com_custos":"Late check-out (com custo)",
    "problemas_com_pagamento":"Problemas com pagamento",
    "detalhes_da_acomodacao":"Dúvidas sobre acomodação",
    "problemas_com_link_de_pagamento":"Problemas com link de pagamento",
    "reembolso":"Reembolso","estacionamento_dificuldades_de_acesso":"Estacionamento — acesso",
    "reclamacao_de_internet":"Reclamação de internet","item_danificado_no_apto":"Item danificado",
    "reclamacao_estadia":"Reclamação de estadia","comercial":"Comercial",
    "atendimento-comercial":"Atendimento comercial","hospedar":"Hospedar","morar":"Morar",
    "contato_ativo":"Contato ativo","falta_de_interacao":"Falta de interação",
    "aviso_de_check_out":"Aviso de check-out","checkout_manual":"Checkout manual",
    "check_in_erro_de_verificacao":"Check-in — erro de verificação",
    "check_in_manual_documento_digital":"Check-in manual — doc digital",
    "suspeita_de_fraude":"Suspeita de fraude",
    "abertura_remota_check_out_vencido":"Abertura remota (check-out vencido)",
    "abertura_remota_dificuldades_do_hospede":"Abertura remota — dificuldades",
    "solicitacao_de_early_check_in_sem_disponibilidade":"Early check-in (sem disponibilidade)",
    "estacionamento_reserva":"Estacionamento — reserva","maleiro":"Maleiro",
    "mensalista_solicitacao_de_link_de_pagamento":"Mensalista — link de pagamento",
    "item_faltante_na_unidade":"Item faltante",
    "solicitacao_de_servico_troca_de_enxoval":"Troca de enxoval",
    "reposicao_de_amenidades":"Reposição de amenidades",
    "reclamacao_de_barulho":"Reclamação de barulho","limpeza_nao_realizada":"Limpeza não realizada",
    "ar_condicionado":"Ar condicionado","hidraulica":"Hidráulica",
    "aviso_de_falta_de_energia":"Falta de energia","feedback_estadia":"Feedback de estadia",
    "mkt_hospede_frequent":"Hóspede frequente","novos_negocios":"Novos negócios",
    "trabalhe_conosco":"Trabalhe conosco","conversa_operacional":"Conversa operacional",
    "teste":"Teste","checkin_b2b":"Check-in B2B","reserva_upgrade":"Upgrade de reserva",
    "cupom_de_desconto_reservas":"Cupom de desconto","mensalista_contrato_venda":"Mensalista — contrato",
    "solicitacao_de_late_check_out_sem_disponibilidade":"Late check-out (sem disponibilidade)",
    "duvidas_omnibees":"Dúvidas Omnibees","alteracao_de_dados_cadastrais":"Alteração de dados cadastrais",
    "comercial_hospedar":"Comercial — hospedar","comercial_morar":"Comercial — morar",
    "erro_de_verificacao_checkin_manual":"Check-in manual — erro de verificação",
    "erro_senha_null":"Erro senha null",
    "eletrica":"Elétrica",
    "mobiliario":"Mobiliário",
    "equipamentos":"Equipamentos",
    "aviso_falta_de_agua":"Falta de água",
    "contato_ativo_comunicacoes":"Contato ativo — comunicações",
    "contato_ativo_cobranca":"Contato ativo — cobrança",
    "contato_ativo_autorizacao_entrada":"Contato ativo — autorização de entrada",
}

# ── Funções auxiliares ───────────────────────────────────────

def lbl(tag):
    t = tag.lower().strip()
    return LABELS.get(t, t.replace("_"," ").replace("-"," ").title())

def _v(val):
    if val is None: return None
    s = str(val).strip()
    return None if s == '' else s

def parse_min(val):
    v = _v(val)
    if not v: return None
    try:
        p = v.split(":")
        if len(p)==3: return int(p[0])*60+int(p[1])+float(p[2])/60
    except: pass
    return None

def parse_dt(val):
    v = _v(val)
    if not v: return None
    for f in ("%d/%m/%Y","%Y-%m-%d"):
        try: return datetime.strptime(v[:10],f).date()
        except: pass
    return None

def sem_lbl(d):
    return (d-timedelta(days=d.weekday())).strftime("%d/%m")

def avg(lst):
    lst = [x for x in lst if x is not None and x>=0]
    return round(sum(lst)/len(lst),1) if lst else None

def detect_emp(tag):
    t = tag.lower()
    for kw, nome in EMPREEND_KEYWORDS:
        if kw in t: return nome
    return None

def classify(tag):
    t = tag.lower().strip()
    cids = set()
    for p, cid in CLASSIF_PREFIXOS:
        if t.startswith(p): cids.add(cid)
    if t in CLASSIF_EXACT:
        val = CLASSIF_EXACT[t]
        if isinstance(val, list): cids.update(val)
        else: cids.add(val)
    return cids

def is_motivo(tag):
    """Retorna True se a tag deve ser contada como motivo de contato."""
    t = tag.lower().strip()
    if t in TAGS_STATUS_CLIENTE: return False
    if t in MAPA_PRIORIDADE: return False
    if t.startswith("hospedado_"): return False  # tags de localização
    if detect_emp(t): return False
    return True

def _safe_float(val):
    v = _v(val)
    if not v: return 0.0
    cleaned = str(v).replace(',','.').replace('%','').replace('R$','').replace('\xa0','').strip()
    try: return float(cleaned)
    except: return 0.0

def _safe_int(val):
    v = _v(val)
    if not v: return 0
    cleaned = str(v).replace(',','.').replace('%','').strip()
    try: return int(float(cleaned))
    except: return 0

def _parse_mes_key(val, ano_base="2026"):
    s = str(val or "").strip().lower()
    if not s: return None
    m = re.match(r'(\d{4})-(\d{2})', s)
    if m: return f"{m.group(1)}-{m.group(2)}"
    m = re.match(r'(\d{1,2})[/-](\d{4})', s)
    if m: return f"{m.group(2)}-{m.group(1).zfill(2)}"
    for nome, num in MES_NUM.items():
        if s.startswith(nome):
            yr_m = re.search(r'\d{2,4}', s[len(nome):])
            yr = yr_m.group() if yr_m else ano_base
            if len(yr) == 2: yr = "20" + yr
            return f"{yr}-{num}"
    return None

def _build_metrics(tids, ticket_base, ticket_tags, occ_lookup=None):
    """Agrega métricas para um conjunto de ticket IDs."""
    cc = defaultdict(int)
    ctpr = defaultdict(list); cttf = defaultdict(list); ctags = defaultdict(Counter)
    emp_c = Counter(); emp_m = defaultdict(Counter); clu_c = Counter()
    ag_c = defaultdict(int); ag_tpr = defaultdict(list)
    ag_ttf = defaultdict(list); ag_msg = defaultdict(list)
    pri = Counter(); tprs = []; ttfs = []
    canal_c = Counter()
    # Cruzamentos — item 9
    manut_emp = Counter(); limp_emp = Counter()
    # Tags em atendimentos de prioridade — item 3
    pri_tag_c = {"Alta": Counter(), "Média": Counter(), "Baixa": Counter()}

    for tid in tids:
        base = ticket_base[tid]
        tags = [t.strip() for t in ticket_tags.get(tid, [])]
        tpr = base["tpr"]; ttf = base["ttf"]
        if tpr is not None: tprs.append(tpr)
        if ttf is not None: ttfs.append(ttf)
        ag = base["agente"] or "Sem agente"
        ag_c[ag]+=1; ag_tpr[ag].append(tpr); ag_ttf[ag].append(ttf); ag_msg[ag].append(base["msgs"])
        canal_c[base.get("canal","—")] += 1

        # Empreendimentos
        emps = set()
        for tag in tags:
            e = detect_emp(tag)
            if e: emps.add(e)
        # Filtrar apenas ativos
        emps = {e for e in emps if e in EMPREEND_ATIVOS}

        for e in emps:
            emp_c[e]+=1; clu_c[CLUSTERS.get(e,"Outros")]+=1

        # Prioridade
        ticket_prio = None
        for tag in tags:
            t = tag.lower().strip()
            if t in MAPA_PRIORIDADE:
                ticket_prio = MAPA_PRIORIDADE[t]
                pri[ticket_prio]+=1
                break

        mot = [t for t in tags if is_motivo(t)]
        for e in emps:
            for mt in mot: emp_m[e][lbl(mt)]+=1

        # Tags por prioridade — item 3
        if ticket_prio:
            for mt in mot:
                pri_tag_c[ticket_prio][lbl(mt)] += 1

        tclassifs = set()
        for tag in mot:
            for cid in classify(tag):
                tclassifs.add(cid); ctags[cid][lbl(tag)]+=1
        for cid in tclassifs:
            cc[cid]+=1; ctpr[cid].append(tpr); cttf[cid].append(ttf)

        # Cruzamentos Manutenção / Limpeza x Empreendimento — item 9
        if "manutencao" in tclassifs:
            for e in emps: manut_emp[e]+=1
            if not emps: manut_emp["Sem empreendimento"]+=1
        if "limpeza" in tclassifs:
            for e in emps: limp_emp[e]+=1
            if not emps: limp_emp["Sem empreendimento"]+=1

    total = len(tids)

    # Classifs com % — itens 1 e 2
    classifs = []
    for cid in CLASSIF_ORDER:
        cnt = cc.get(cid, 0)
        pct = round(cnt / total * 100, 1) if total > 0 else 0
        classifs.append({
            "id": cid, "label": CLASSIF_META[cid]["label"], "cor": CLASSIF_META[cid]["cor"],
            "count": cnt, "pct": pct,
            "tpr_med": avg(ctpr.get(cid,[])), "ttf_med": avg(cttf.get(cid,[])),
            "top_tags": [{"label":k,"count":v} for k,v in ctags[cid].most_common(8)],
        })

    emps_lista = []
    for nome in EMPREEND_ORDER:
        tickets = emp_c.get(nome, 0)
        occ = (occ_lookup or {}).get(nome, {})
        reservas = occ.get("reservas", 0)
        occ_pct = occ.get("occ_pct", None)
        un_total = occ.get("un_total", 0)
        taxa = round(tickets / reservas * 100, 1) if reservas > 0 else None
        pct_emp = round(tickets / total * 100, 1) if total > 0 else 0
        top8 = [{"label":k,"count":v} for k,v in emp_m[nome].most_common(8)]
        manut_cnt = manut_emp.get(nome, 0)
        limp_cnt  = limp_emp.get(nome, 0)
        emps_lista.append({
            "nome": nome, "cluster": CLUSTERS.get(nome,"Outros"),
            "count": tickets, "pct": pct_emp, "top_motivos": top8,
            "reservas": reservas, "occ_pct": occ_pct, "un_total": un_total,
            "taxa_contato": taxa,
            "manutencao": manut_cnt, "limpeza": limp_cnt,
        })
    emps_lista.sort(key=lambda x: -x["count"])

    agentes = [{"nome":ag,"tickets":cnt,"pct":round(cnt/total*100,1) if total>0 else 0,
        "tpr_med":avg(ag_tpr[ag]),"ttf_med":avg(ag_ttf[ag]),"msgs_med":avg(ag_msg[ag]),
        "nps": None, "aderencia": None}  # Colunas futuras — configurar na planilha
        for ag,cnt in sorted(ag_c.items(),key=lambda x:-x[1])]

    clusters_lista = [{"nome":cl,"count":clu_c.get(cl,0)} for cl in CLUSTER_ORDER]

    canais_lista = [{"canal":k,"count":v} for k,v in canal_c.most_common()]

    # Tags de prioridade — item 3
    priority_tags = {
        "Alta":  [{"label":k,"count":v} for k,v in pri_tag_c["Alta"].most_common(10)],
        "Média": [{"label":k,"count":v} for k,v in pri_tag_c["Média"].most_common(10)],
        "Baixa": [{"label":k,"count":v} for k,v in pri_tag_c["Baixa"].most_common(10)],
    }

    return {
        "total": total,
        "tpr_med": avg(tprs),
        "ttf_med": avg(ttfs),
        "classifs": classifs,
        "empreendimentos": emps_lista,
        "agentes": agentes,
        "clusters": clusters_lista,
        "canais": canais_lista,
        "prioridades": {
            "Alta":  pri.get("Alta",0),
            "Media": pri.get("Média",0),
            "Baixa": pri.get("Baixa",0),
        },
        "priority_tags": priority_tags,
    }

def _volume_insights(vd):
    """Calcula insights sobre volume diário — item 13."""
    if not vd:
        return {}
    sorted_days = sorted(vd.items())
    counts = [c for _, c in sorted_days]
    if not counts:
        return {}
    total_c = sum(counts)
    avg_daily = round(total_c / len(counts), 1)
    max_day = max(sorted_days, key=lambda x: x[1])
    min_day = min(sorted_days, key=lambda x: x[1])
    # Tendência: últimos 7 dias vs 7 dias anteriores
    trend_pct = None
    trend_dir = "estável"
    if len(counts) >= 14:
        last7 = sum(counts[-7:])
        prev7 = sum(counts[-14:-7])
        if prev7 > 0:
            trend_pct = round((last7 - prev7) / prev7 * 100, 1)
            if trend_pct > 5: trend_dir = "crescimento"
            elif trend_pct < -5: trend_dir = "redução"
    return {
        "max_day": {
            "date": max_day[0],
            "date_fmt": max_day[0][8:]+"/"+max_day[0][5:7],
            "count": max_day[1],
        },
        "min_day": {
            "date": min_day[0],
            "date_fmt": min_day[0][8:]+"/"+min_day[0][5:7],
            "count": min_day[1],
        },
        "avg_daily": avg_daily,
        "total_dias": len(counts),
        "trend_pct": trend_pct,
        "trend_dir": trend_dir,
    }

def _tags_pendentes(all_tags_counter):
    """
    Retorna tags que não pertencem a nenhuma categoria — item 10.
    Exclui: status de cliente, prioridade, empreendimentos, tags hospedado_.
    """
    pendentes = []
    for tag, count in all_tags_counter.most_common():
        t = tag.lower().strip()
        if t in TAGS_STATUS_CLIENTE: continue
        if t in MAPA_PRIORIDADE: continue
        if t.startswith("hospedado_"): continue
        if detect_emp(t): continue
        cids = classify(t)
        if cids: continue
        pendentes.append({"tag": tag, "count": count, "label": lbl(tag)})
    return pendentes

# ── Função principal ─────────────────────────────────────────

def processar(rows_droz, rows_occ=None):
    """
    Processa dados de atendimento e OCC.
    rows_droz : lista de linhas com dados de tickets (sem header)
    rows_occ  : lista de linhas com dados de OCC (sem header), ou None
    Retorna: payload dict pronto para JSON
    """
    # 1. Leitura de tickets
    ticket_base = {}
    ticket_tags = defaultdict(list)
    all_tags_counter = Counter()

    for row in rows_droz:
        row = list(row) + [None]*30
        tid = _v(row[1])
        if not tid: continue
        tag = _v(row[23])
        if tid not in ticket_base:
            ticket_base[tid] = {
                "agente": _v(row[19]),
                "data":   parse_dt(row[22]),
                "tpr":    parse_min(row[12]),
                "ttf":    parse_min(row[13]),
                "msgs":   _safe_int(row[14]),
                "canal":  _v(row[18]) or "—",
                # Colunas opcionais: NPS (col 24), Pesquisas Respondidas (col 25)
                # Adicione essas colunas na planilha para habilitar os indicadores
                "nps":       _safe_float(row[24]) if _v(row[24]) else None,
                "pesquisa":  _safe_int(row[25]) if _v(row[25]) else None,
            }
        if tag:
            ticket_tags[tid].append(tag.strip())
            all_tags_counter[tag.strip().lower()] += 1

    # 2. Volume por período
    vd = Counter(); vs = Counter(); vm = Counter()
    for tid, base in ticket_base.items():
        dt = base["data"]
        if dt:
            vd[dt.strftime("%Y-%m-%d")]+=1
            vm[dt.strftime("%Y-%m")]+=1
            vs[sem_lbl(dt)]+=1

    # 3. Leitura de OCC — item 14
    # Estrutura atual: col0=Empreendimento, col1=TotalUnidades, col2=UnDisp,
    #                  col3=UnOcupadas, col4=DM, col5=Faturamento, col6=Ocupação%
    occ_agg = {}
    if rows_occ:
        for row in rows_occ:
            row = list(row) + [None]*9
            emp_raw = str(_v(row[0]) or "").strip()
            if not emp_raw: continue
            canon = detect_emp(emp_raw)
            if not canon: continue
            if canon not in EMPREEND_ATIVOS: continue
            un_total = _safe_int(row[1])
            un_ocup  = _safe_int(row[3])
            occ_r    = _safe_float(row[6])
            occ_pct  = round(occ_r*100, 1) if 0 < occ_r <= 1.0 else round(occ_r, 1)
            fat      = _safe_float(row[5])
            occ_agg[canon] = {
                "reservas":  un_ocup,
                "occ_pct":   occ_pct if occ_r > 0 else None,
                "faturamento": fat,
                "un_total":  un_total,
            }

    # 4. Métricas globais
    all_tids = list(ticket_base.keys())
    gm = _build_metrics(all_tids, ticket_base, ticket_tags, occ_agg)

    # 5. Volume charts
    dias_s = sorted(vd.items())
    seen_w = []; seen_ws = set()
    for d_str, _ in dias_s:
        d = datetime.strptime(d_str,"%Y-%m-%d").date()
        lw = sem_lbl(d)
        if lw not in seen_ws: seen_ws.add(lw); seen_w.append(lw)
    sems = [{"d":w,"c":vs[w]} for w in seen_w]
    meses_vol = [{"d":NM.get(m[5:7],m[5:7]),"c":c,"k":m} for m,c in sorted(vm.items())]

    # 6. Métricas por mês
    por_mes = {}
    for mes_key in sorted(vm.keys()):
        month_tids = [tid for tid, base in ticket_base.items()
                      if base["data"] and base["data"].strftime("%Y-%m") == mes_key]
        mm = _build_metrics(month_tids, ticket_base, ticket_tags, occ_agg)
        month_dias = [{"d":d[8:]+"/"+d[5:7],"c":c}
                      for d,c in sorted(vd.items()) if d[:7]==mes_key]
        month_vd = Counter({d: c for d, c in vd.items() if d[:7] == mes_key})
        por_mes[mes_key] = {
            **mm, "label": NM.get(mes_key[5:7], mes_key),
            "dias": month_dias,
            "volume_insights": _volume_insights(month_vd),
        }

    # 7. Tags pendentes de categorização — item 10
    tags_pendentes = _tags_pendentes(all_tags_counter)

    # 8. Volume insights globais — item 13
    vol_insights = _volume_insights(vd)

    # 9. Período label
    pi = sorted(vm.keys())[0] if vm else ""; pf = sorted(vm.keys())[-1] if vm else ""
    if pi == pf:
        periodo = NM.get(pi[5:7],"") + " " + pi[:4] if pi else ""
    else:
        periodo = NM.get(pi[5:7],"")+" – "+NM.get(pf[5:7],"")+" "+pf[:4] if pi else ""

    # 10. OCC payload para visualização — item 14
    occ_lista = [
        {"nome":nome, **vals}
        for nome, vals in occ_agg.items()
    ]

    return {
        "gerado_em": datetime.now(TZ_BR).strftime("%d/%m/%Y %H:%M"),  # item 12
        "periodo": periodo,
        "total": gm["total"],
        "tpr_med": gm["tpr_med"],
        "ttf_med": gm["ttf_med"],
        "dias": [{"d":d,"c":c} for d,c in dias_s],
        "semanas": sems,
        "meses": meses_vol,
        "classifs": gm["classifs"],
        "empreendimentos": gm["empreendimentos"],
        "clusters": gm["clusters"],
        "canais": gm["canais"],
        "prioridades": gm["prioridades"],
        "priority_tags": gm["priority_tags"],
        "agentes": gm["agentes"],
        "correl": sorted(
            [e for e in gm["empreendimentos"] if e["taxa_contato"] is not None],
            key=lambda x: -x["taxa_contato"]
        ),
        "occ": occ_lista,
        "por_mes": por_mes,
        "tags_pendentes": tags_pendentes,
        "volume_insights": vol_insights,
    }

# ── Uso local ────────────────────────────────────────────────

if __name__ == "__main__":
    import openpyxl
    print("Carregando planilha...")
    wb = openpyxl.load_workbook(ARQUIVO_ENTRADA)

    sheet_droz = None
    for candidate in ["Atendimentos","Input Droz (Abr-Mai)","Resultado da consulta"]:
        if candidate in wb.sheetnames:
            sheet_droz = wb[candidate]; break
    if not sheet_droz:
        sheet_droz = wb.active
    print("Aba de atendimentos:", sheet_droz.title)

    ws_occ = wb["Reservas e OCC"] if "Reservas e OCC" in wb.sheetnames else None
    print("Aba de OCC:", ws_occ.title if ws_occ else "não encontrada")

    rows_droz = list(sheet_droz.iter_rows(min_row=2, values_only=True))
    rows_occ  = list(ws_occ.iter_rows(min_row=2, values_only=True)) if ws_occ else None

    payload = processar(rows_droz, rows_occ)
    print("Tickets únicos:", payload["total"])
    print("TPR med:", payload["tpr_med"], "| TTF med:", payload["ttf_med"])
    print("Empreendimentos com dados:", [e["nome"] for e in payload["empreendimentos"] if e["count"]>0])
    print("Tags pendentes:", len(payload["tags_pendentes"]))
    print("Volume insights:", payload["volume_insights"])

    with open(ARQUIVO_TEMPLATE,"r",encoding="utf-8") as f:
        html = f.read()
    html = html.replace("__DATA__", json.dumps(payload, ensure_ascii=False))
    with open(ARQUIVO_SAIDA,"w",encoding="utf-8") as f:
        f.write(html)
    print("Dashboard gerado:", ARQUIVO_SAIDA)
